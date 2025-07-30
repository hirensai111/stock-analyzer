"""
Enhanced Excel report generator for Stock Analyzer.
Creates professional Excel reports with charts, visualizations, and comprehensive analysis.
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, DataBarRule
from openpyxl.chart import LineChart, Reference, BarChart, StockChart, Series
from openpyxl.chart.axis import DateAxis
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.cell.cell import MergedCell
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.figure import Figure
from matplotlib.backends.backend_agg import FigureCanvasAgg

from openpyxl.drawing.image import Image
import io
from typing import Dict, Any

from news_integration_bridge import news_bridge
from event_analyzer import enhanced_analyzer, PriceEvent, EventAnalysis

from config import config
from utils import get_logger, DataFormatter

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class ExcelReportGenerator:
    """Generate professional Excel reports for stock analysis with charts and styling."""
    
    def __init__(self):
        self.logger = get_logger("ExcelReportGenerator")
        self.formatter = DataFormatter()
        self._setup_styles()
        
        self.enhanced_analyzer = enhanced_analyzer
        
        self.significance_threshold = 3.0  # For identifying significant price movements
        self.max_days_to_analyze = 10
        
    def _setup_styles(self):
        """Setup professional styles for the Excel report."""
        # Color scheme
        self.colors = {
            'primary': 'FF1F4788',      # Dark blue
            'secondary': 'FF2E86AB',    # Medium blue
            'accent': 'FF5EB3D6',       # Light blue
            'success': 'FF27AE60',      # Green
            'warning': 'FFF39C12',      # Orange
            'danger': 'FFE74C3C',       # Red
            'light': 'FFF8F9FA',        # Light gray
            'dark': 'FF2C3E50',          # Dark gray
            'event_bullish': 'FFC6EFCE',  # Light green for bullish events
            'event_bearish': 'FFFFC7CE',  # Light red for bearish events
        }
        
        # Create named styles
        self.styles = {}
        
        # Header style
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color="FFFFFFFF", size=12)
        header_style.fill = PatternFill(start_color=self.colors['primary'], 
                                       end_color=self.colors['primary'], 
                                       fill_type="solid")
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thick')
        )
        self.styles['header'] = header_style
        
        # Subheader style
        subheader_style = NamedStyle(name="subheader_style")
        subheader_style.font = Font(bold=True, color="FFFFFFFF", size=11)
        subheader_style.fill = PatternFill(start_color=self.colors['secondary'], 
                                          end_color=self.colors['secondary'], 
                                          fill_type="solid")
        subheader_style.alignment = Alignment(horizontal="left", vertical="center")
        self.styles['subheader'] = subheader_style
        
        # Title style
        title_style = NamedStyle(name="title_style")
        title_style.font = Font(bold=True, size=16, color=self.colors['primary'])
        title_style.alignment = Alignment(horizontal="center", vertical="center")
        self.styles['title'] = title_style
        
        # Number format styles
        self.number_formats = {
            'currency': '$#,##0.00',
            'percentage': '0.00%',
            'number': '#,##0.00',
            'integer': '#,##0',
            'date': 'yyyy-mm-dd'
        }
        
    def generate_report(self, stock_data: Dict[str, Any]) -> str:
        """
        Generate comprehensive Excel report with charts and visualizations.
        
        Args:
            stock_data: Dictionary containing all stock analysis data
            
        Returns:
            Path to generated Excel file
        """
        ticker = stock_data['ticker']
        self.logger.info(f"Generating enhanced Excel report for {ticker}")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # ANALYZE PRICE EVENTS BEFORE CREATING SHEETS
        tech_data = stock_data.get('technical_data', stock_data.get('raw_data', pd.DataFrame()))
        if not tech_data.empty:
            analyzed_data = self._analyze_price_events(tech_data, ticker)
            stock_data['event_analysis'] = analyzed_data  # Add analyzed data back to stock_data
            stock_data['technical_data'] = analyzed_data
        
        # Create sheets in specific order
        self._create_summary_sheet(wb, stock_data)
        self._create_company_info_sheet(wb, stock_data)
        self._create_price_chart_sheet(wb, stock_data)
        self._create_technical_analysis_sheet(wb, stock_data)
        self._create_sentiment_analysis_sheet(wb, stock_data)
        self._create_performance_metrics_sheet(wb, stock_data)
        self._create_raw_data_sheet(wb, stock_data)
        self._create_data_quality_sheet(wb, stock_data)
        self._create_metadata_sheet(wb, stock_data)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{ticker}_analysis_report_{timestamp}.xlsx"
        filepath = config.OUTPUT_DIR / filename
        
        # Save workbook
        wb.save(filepath)
        self.logger.info(f"Enhanced Excel report saved to {filepath}")
        
        return str(filepath)
    
    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create executive summary sheet with key metrics and signals."""
        ws = wb.create_sheet("Summary")
        ticker = data['ticker']
        summary = data['summary_statistics']
        company_info = data['company_info']
        
        # Add title
        ws['A1'] = f"{ticker} - Stock Analysis Summary"
        ws['A1'].style = self.styles['title']
        ws.merge_cells('A1:H1')
        
        # Add generation date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:H2')
        
        # Company Overview Section
        row = 4
        ws[f'A{row}'] = "Company Overview"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        company_data = [
            ['Company Name:', company_info.get('longName', 'N/A')],
            ['Sector:', company_info.get('sector', 'N/A')],
            ['Industry:', company_info.get('industry', 'N/A')],
            ['Market Cap:', self.formatter.format_market_cap(company_info.get('marketCap', 0))],
            ['Exchange:', company_info.get('exchange', 'N/A')]
        ]
        
        for label, value in company_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:C{row}')
            row += 1
        
        # Price Information Section
        row += 1
        ws[f'A{row}'] = "Price Information"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        price_data = [
            ['Current Price:', f"${summary['current_price']:.2f}"],
            ['Day Change:', f"${summary['price_change_1d']:.2f} ({summary['price_change_1d_pct']:.2f}%)"],
            ['52-Week Range:', f"${summary['52_week_low']:.2f} - ${summary['52_week_high']:.2f}"],
            ['Volume:', f"{summary['volume']:,}"],
            ['Avg Volume:', f"{summary['avg_volume']:,}"]
        ]
        
        for label, value in price_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:C{row}')
            
            # Add conditional formatting for day change
            if label == 'Day Change:':
                if summary['price_change_1d'] > 0:
                    ws[f'B{row}'].font = Font(color=self.colors['success'])
                elif summary['price_change_1d'] < 0:
                    ws[f'B{row}'].font = Font(color=self.colors['danger'])
            row += 1
        
        # Trading Signals Section
        row += 1
        ws[f'A{row}'] = "Trading Signals"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        signals = summary.get('signals', {})
        signal_data = [
            ['Overall Signal:', signals.get('overall', 'Neutral')],
            ['Trend:', signals.get('trend', 'Neutral')],
            ['Momentum:', signals.get('momentum', 'Neutral')],
            ['Volume:', signals.get('volume', 'Normal')]
        ]
        
        for label, value in signal_data:
            ws[f'E{row}'] = label
            ws[f'E{row}'].font = Font(bold=True)
            ws[f'F{row}'] = value
            ws.merge_cells(f'F{row}:G{row}')
            
            # Color code signals
            if value == 'Bullish' or value == 'Buy':
                ws[f'F{row}'].font = Font(color=self.colors['success'], bold=True)
            elif value == 'Bearish' or value == 'Sell':
                ws[f'F{row}'].font = Font(color=self.colors['danger'], bold=True)
            elif value == 'High':
                ws[f'F{row}'].font = Font(color=self.colors['warning'], bold=True)
            row += 1
        
        # Performance Returns Section
        row = row - 4  # Align with signals
        ws[f'A{row}'] = "Performance Returns"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        returns = summary.get('returns', {})
        return_periods = [
            ('1 Day', returns.get('1_day')),
            ('1 Week', returns.get('1_week')),
            ('1 Month', returns.get('1_month')),
            ('3 Months', returns.get('3_months')),
            ('6 Months', returns.get('6_months')),
            ('1 Year', returns.get('1_year'))
        ]
        
        for period, value in return_periods:
            if value is not None:
                ws[f'A{row}'] = f"{period}:"
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = f"{value:.2f}%"
                
                # Color code returns
                if value > 0:
                    ws[f'B{row}'].font = Font(color=self.colors['success'])
                elif value < 0:
                    ws[f'B{row}'].font = Font(color=self.colors['danger'])
                row += 1
        
        # Risk Metrics Section
        row += 2
        ws[f'A{row}'] = "Risk Metrics"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        risk_data = [
            ['Annual Volatility:', f"{summary.get('volatility_annual', 0):.2f}%"],
            ['Sharpe Ratio:', f"{summary.get('sharpe_ratio', 0):.2f}"],
            ['Maximum Drawdown:', f"{summary.get('max_drawdown', 0):.2f}%"],
            ['Beta (vs S&P 500):', f"{summary.get('beta', 'N/A')}"]
        ]
        
        for label, value in risk_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:C{row}')
            row += 1
        
        # Technical Indicators Section
        ws[f'E{row-4}'] = "Technical Indicators"
        ws[f'E{row-4}'].style = self.styles['subheader']
        ws.merge_cells(f'E{row-4}:H{row-4}')
        
        indicators = summary.get('technical_indicators', {})
        indicator_data = [
            ['RSI (14):', f"{indicators.get('rsi', 0):.2f}"],
            ['MACD:', f"{indicators.get('macd', 0):.4f}"],
            ['SMA 20:', f"${indicators.get('sma_20', 0):.2f}"],
            ['SMA 50:', f"${indicators.get('sma_50', 0):.2f}"],
            ['SMA 200:', f"${indicators.get('sma_200', 0):.2f}"]
        ]
        
        row = row - 3
        for label, value in indicator_data:
            ws[f'E{row}'] = label
            ws[f'E{row}'].font = Font(bold=True)
            ws[f'F{row}'] = value
            ws.merge_cells(f'F{row}:G{row}')
            row += 1
            
        # Recent Significant Events Section (if available)
        if 'event_analysis' in data:
            event_data = data['event_analysis']
            significant_events = event_data[event_data.get('Is_Significant', False)]
            recent_events = significant_events.tail(5)  # Last 5 significant events
            
            if not recent_events.empty:
                row += 2
                ws[f'A{row}'] = "Recent Significant Events"
                ws[f'A{row}'].style = self.styles['subheader']
                ws.merge_cells(f'A{row}:H{row}')
                
                row += 1
                for idx, event in recent_events.iterrows():
                    date_str = idx.strftime('%Y-%m-%d') if hasattr(idx, 'strftime') else str(idx)
                    change = event.get('Daily_Change_%', 0)
                    event_type = event.get('Event_Type', 'Unknown')
                    
                    ws[f'A{row}'] = f"{date_str}: {change:+.2f}% - {event_type}"
                    ws.merge_cells(f'A{row}:H{row}')
                    
                    if change > 0:
                        ws[f'A{row}'].font = Font(color=self.colors['success'])
                    else:
                        ws[f'A{row}'].font = Font(color=self.colors['danger'])
                    
                    row += 1
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            # Find the first non-merged cell to get the column letter
            column_letter = None
            for cell in column_cells:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break
            if not column_letter:
                continue  # skip if all cells are merged

            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_company_info_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create detailed company information sheet."""
        ws = wb.create_sheet("Company Info")
        ticker = data['ticker']
        company_info = data['company_info']
        
        # Title
        ws['A1'] = f"{ticker} - Company Information"
        ws['A1'].style = self.styles['title']
        ws.merge_cells('A1:F1')
        
        # Basic Information
        row = 3
        ws[f'A{row}'] = "Basic Information"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        basic_info = [
            ['Company Name', company_info.get('longName', 'N/A')],
            ['Symbol', ticker],
            ['Exchange', company_info.get('exchange', 'N/A')],
            ['Currency', company_info.get('currency', 'USD')],
            ['Country', company_info.get('country', 'N/A')],
            ['Website', company_info.get('website', 'N/A')],
            ['Employees', f"{company_info.get('fullTimeEmployees', 'N/A'):,}" if isinstance(company_info.get('fullTimeEmployees'), int) else 'N/A']
        ]
        
        for label, value in basic_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:D{row}')
            row += 1
        
        # Business Description
        row += 1
        ws[f'A{row}'] = "Business Description"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        description = company_info.get('longBusinessSummary', 'No description available')
        ws[f'A{row}'] = description
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:F{row+5}')
        ws.row_dimensions[row].height = 100
        
        # Financial Metrics
        row += 7
        ws[f'A{row}'] = "Financial Metrics"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        financial_data = [
            ['Market Cap', self.formatter.format_market_cap(company_info.get('marketCap', 0))],
            ['Enterprise Value', self.formatter.format_market_cap(company_info.get('enterpriseValue', 0))],
            ['Revenue (TTM)', self.formatter.format_market_cap(company_info.get('totalRevenue', 0))],
            ['Gross Profit', self.formatter.format_market_cap(company_info.get('grossProfits', 0))],
            ['EBITDA', self.formatter.format_market_cap(company_info.get('ebitda', 0))],
            ['Free Cash Flow', self.formatter.format_market_cap(company_info.get('freeCashflow', 0))]
        ]
        
        for i, (label, value) in enumerate(financial_data):
            col = 'A' if i % 2 == 0 else 'D'
            value_col = 'B' if i % 2 == 0 else 'E'
            
            ws[f'{col}{row}'] = label
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{value_col}{row}'] = value
            
            if i % 2 == 1:
                row += 1
        
        # Valuation Ratios
        row += 2
        ws[f'A{row}'] = "Valuation Ratios"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        valuation_data = [
            ['P/E Ratio', f"{company_info.get('trailingPE', 0):.2f}" if company_info.get('trailingPE') else 'N/A'],
            ['Forward P/E', f"{company_info.get('forwardPE', 0):.2f}" if company_info.get('forwardPE') else 'N/A'],
            ['PEG Ratio', f"{company_info.get('pegRatio', 0):.2f}" if company_info.get('pegRatio') else 'N/A'],
            ['Price/Sales', f"{company_info.get('priceToSalesTrailing12Months', 0):.2f}" if company_info.get('priceToSalesTrailing12Months') else 'N/A'],
            ['Price/Book', f"{company_info.get('priceToBook', 0):.2f}" if company_info.get('priceToBook') else 'N/A'],
            ['EV/EBITDA', f"{company_info.get('enterpriseToEbitda', 0):.2f}" if company_info.get('enterpriseToEbitda') else 'N/A']
        ]
        
        for i, (label, value) in enumerate(valuation_data):
            col = 'A' if i % 2 == 0 else 'D'
            value_col = 'B' if i % 2 == 0 else 'E'
            
            ws[f'{col}{row}'] = label
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{value_col}{row}'] = value
            
            if i % 2 == 1:
                row += 1
        
        # Dividend Information
        row += 2
        ws[f'A{row}'] = "Dividend Information"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        dividend_data = [
            ['Dividend Rate', f"${company_info.get('dividendRate', 0):.2f}" if company_info.get('dividendRate') else 'N/A'],
            ['Dividend Yield', f"{company_info.get('dividendYield', 0)*100:.2f}%" if company_info.get('dividendYield') else 'N/A'],
            ['Ex-Dividend Date', datetime.fromtimestamp(company_info.get('exDividendDate', 0)).replace(tzinfo=None).strftime('%Y-%m-%d') if company_info.get('exDividendDate') else 'N/A'],
            ['Payout Ratio', f"{company_info.get('payoutRatio', 0)*100:.2f}%" if company_info.get('payoutRatio') else 'N/A']
        ]
        
        for label, value in dividend_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:C{row}')
            row += 1
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            # Find the first non-merged cell to get the column letter
            column_letter = None
            for cell in column_cells:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break
            if not column_letter:
                continue  # skip if all cells are merged

            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_price_chart_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create sheet with price charts and volume visualization using matplotlib."""
        ws = wb.create_sheet("Price Charts")
        ticker = data['ticker']
        
        # Title
        ws['A1'] = f"{ticker} - Price & Volume Charts"
        ws['A1'].style = self.styles['title']
        ws.merge_cells('A1:J1')
        
        # Get technical data
        tech_data = data.get('technical_data', data.get('raw_data'))
        
        if not isinstance(tech_data, pd.DataFrame) or tech_data.empty:
            ws['A3'] = "No data available for charts"
            return
        
        # Prepare data for charts (last 252 trading days)
        chart_data = tech_data.dropna(subset=["Close"]).tail(252).copy()
        
        # Set consistent style for all charts
        plt.style.use('seaborn-v0_8-darkgrid')
        
        # Create figure with subplots
        fig = plt.figure(figsize=(16, 20))
        
        # 1. Price and Moving Averages Chart
        ax1 = plt.subplot(4, 1, 1)
        ax1.plot(chart_data.index, chart_data['Close'], label='Close Price', color='#0070C0', linewidth=2)
        
        # Add moving averages if they exist
        if 'SMA_20' in chart_data.columns:
            ax1.plot(chart_data.index, chart_data['SMA_20'], label='SMA 20', color='#F39C12', linewidth=1.5, alpha=0.8)
        if 'SMA_50' in chart_data.columns:
            ax1.plot(chart_data.index, chart_data['SMA_50'], label='SMA 50', color='#3498DB', linewidth=1.5, alpha=0.8)
        if 'SMA_200' in chart_data.columns:
            ax1.plot(chart_data.index, chart_data['SMA_200'], label='SMA 200', color='#9B59B6', linewidth=1.5, alpha=0.8)
        
        ax1.set_title(f'{ticker} Price & Moving Averages', fontsize=14, fontweight='bold')
        ax1.set_ylabel('Price ($)', fontsize=12)
        ax1.legend(loc='best')
        ax1.grid(True, alpha=0.3)
        
        # Format x-axis dates
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        ax1.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45)
        
        # 2. Volume Chart
        ax2 = plt.subplot(4, 1, 2)
        colors = ['g' if chart_data['Close'].iloc[i] >= chart_data['Close'].iloc[i-1] else 'r' 
                for i in range(1, len(chart_data))]
        colors.insert(0, 'g')  # First bar green by default
        
        ax2.bar(chart_data.index, chart_data['Volume'], color=colors, alpha=0.7)
        ax2.set_title('Trading Volume', fontsize=14, fontweight='bold')
        ax2.set_ylabel('Volume', fontsize=12)
        ax2.grid(True, alpha=0.3)
        
        # Format volume with scientific notation if needed
        ax2.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1e6:.1f}M' if x >= 1e6 else f'{x/1e3:.0f}K'))
        
        # Format x-axis dates
        ax2.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        ax2.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45)
        
        # 3. RSI Chart
        ax3 = plt.subplot(4, 1, 3)
        if 'RSI' in chart_data.columns and not chart_data['RSI'].isna().all():
            ax3.plot(chart_data.index, chart_data['RSI'], label='RSI', color='#1F77B4', linewidth=2)
            ax3.axhline(y=70, color='r', linestyle='--', alpha=0.7, label='Overbought (70)')
            ax3.axhline(y=30, color='g', linestyle='--', alpha=0.7, label='Oversold (30)')
            ax3.set_ylim(0, 100)
        else:
            ax3.text(0.5, 0.5, 'RSI Data Not Available', transform=ax3.transAxes, 
                    ha='center', va='center', fontsize=12)
        
        ax3.set_title('Relative Strength Index (RSI)', fontsize=14, fontweight='bold')
        ax3.set_ylabel('RSI', fontsize=12)
        ax3.legend(loc='best')
        ax3.grid(True, alpha=0.3)
        
        # Format x-axis dates
        ax3.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        ax3.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        plt.setp(ax3.xaxis.get_majorticklabels(), rotation=45)
        
        # 4. MACD Chart
        ax4 = plt.subplot(4, 1, 4)
        if 'MACD' in chart_data.columns and 'MACD_Signal' in chart_data.columns:
            if not chart_data['MACD'].isna().all():
                ax4.plot(chart_data.index, chart_data['MACD'], label='MACD', color='#0070C0', linewidth=2)
                ax4.plot(chart_data.index, chart_data['MACD_Signal'], label='Signal', color='#ED7D31', linewidth=2)
                
                # Add histogram if available
                if 'MACD_Histogram' in chart_data.columns:
                    colors = ['g' if x > 0 else 'r' for x in chart_data['MACD_Histogram']]
                    ax4.bar(chart_data.index, chart_data['MACD_Histogram'], color=colors, alpha=0.3, label='Histogram')
                
                ax4.axhline(y=0, color='black', linestyle='-', alpha=0.3)
            else:
                ax4.text(0.5, 0.5, 'MACD Data Not Available', transform=ax4.transAxes, 
                        ha='center', va='center', fontsize=12)
        else:
            ax4.text(0.5, 0.5, 'MACD Data Not Available', transform=ax4.transAxes, 
                    ha='center', va='center', fontsize=12)
        
        ax4.set_title('MACD', fontsize=14, fontweight='bold')
        ax4.set_ylabel('MACD', fontsize=12)
        ax4.set_xlabel('Date', fontsize=12)
        ax4.legend(loc='best')
        ax4.grid(True, alpha=0.3)
        
        # Format x-axis dates
        ax4.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        ax4.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        plt.setp(ax4.xaxis.get_majorticklabels(), rotation=45)
        
        # Adjust layout
        plt.tight_layout()
        
        # Save to BytesIO object
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        plt.close(fig)
        
        # Reset buffer position
        img_buffer.seek(0)
        
        # Insert image into Excel
        img = Image(img_buffer)
        img.width = 1200  # Adjust size as needed
        img.height = 1500
        ws.add_image(img, 'A3')
        
        # Add a note about the charts
        ws['A80'] = "Note: Charts are generated as images for better compatibility and reliability."
        ws['A81'] = f"Data range: {chart_data.index[0].strftime('%Y-%m-%d')} to {chart_data.index[-1].strftime('%Y-%m-%d')}"
        
        # Optionally, also save individual chart images
        if hasattr(self, 'output_dir') and self.output_dir:
            self._save_individual_charts(chart_data, ticker)

    def _save_individual_charts(self, chart_data: pd.DataFrame, ticker: str):
        """Save individual chart images to files."""
        import os
        
        charts_dir = os.path.join(self.output_dir, f"{ticker}_charts")
        os.makedirs(charts_dir, exist_ok=True)
        
        # Price chart
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.plot(chart_data.index, chart_data['Close'], label='Close Price', color='#0070C0', linewidth=2)
        if 'SMA_20' in chart_data.columns:
            ax.plot(chart_data.index, chart_data['SMA_20'], label='SMA 20', color='#F39C12', linewidth=1.5)
        if 'SMA_50' in chart_data.columns:
            ax.plot(chart_data.index, chart_data['SMA_50'], label='SMA 50', color='#3498DB', linewidth=1.5)
        if 'SMA_200' in chart_data.columns:
            ax.plot(chart_data.index, chart_data['SMA_200'], label='SMA 200', color='#9B59B6', linewidth=1.5)
        ax.set_title(f'{ticker} Price & Moving Averages')
        ax.set_ylabel('Price ($)')
        ax.legend()
        ax.grid(True, alpha=0.3)
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.savefig(os.path.join(charts_dir, f'{ticker}_price_chart.png'), dpi=150)
        plt.close()
        
        # Volume chart
        fig, ax = plt.subplots(figsize=(10, 4))
        colors = ['g' if chart_data['Close'].iloc[i] >= chart_data['Close'].iloc[i-1] else 'r' 
                for i in range(1, len(chart_data))]
        colors.insert(0, 'g')
        ax.bar(chart_data.index, chart_data['Volume'], color=colors, alpha=0.7)
        ax.set_title(f'{ticker} Trading Volume')
        ax.set_ylabel('Volume')
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{x/1e6:.1f}M' if x >= 1e6 else f'{x/1e3:.0f}K'))
        ax.grid(True, alpha=0.3)
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.savefig(os.path.join(charts_dir, f'{ticker}_volume_chart.png'), dpi=150)
        plt.close()
        
        # RSI chart
        if 'RSI' in chart_data.columns and not chart_data['RSI'].isna().all():
            fig, ax = plt.subplots(figsize=(10, 4))
            ax.plot(chart_data.index, chart_data['RSI'], label='RSI', color='#1F77B4', linewidth=2)
            ax.axhline(y=70, color='r', linestyle='--', alpha=0.7, label='Overbought')
            ax.axhline(y=30, color='g', linestyle='--', alpha=0.7, label='Oversold')
            ax.set_ylim(0, 100)
            ax.set_title(f'{ticker} RSI')
            ax.set_ylabel('RSI')
            ax.legend()
            ax.grid(True, alpha=0.3)
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.savefig(os.path.join(charts_dir, f'{ticker}_rsi_chart.png'), dpi=150)
            plt.close()
        
        # MACD chart
        if 'MACD' in chart_data.columns and not chart_data['MACD'].isna().all():
            fig, ax = plt.subplots(figsize=(10, 4))
            ax.plot(chart_data.index, chart_data['MACD'], label='MACD', color='#0070C0', linewidth=2)
            ax.plot(chart_data.index, chart_data['MACD_Signal'], label='Signal', color='#ED7D31', linewidth=2)
            if 'MACD_Histogram' in chart_data.columns:
                colors = ['g' if x > 0 else 'r' for x in chart_data['MACD_Histogram']]
                ax.bar(chart_data.index, chart_data['MACD_Histogram'], color=colors, alpha=0.3, label='Histogram')
            ax.axhline(y=0, color='black', linestyle='-', alpha=0.3)
            ax.set_title(f'{ticker} MACD')
            ax.set_ylabel('MACD')
            ax.legend()
            ax.grid(True, alpha=0.3)
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.savefig(os.path.join(charts_dir, f'{ticker}_macd_chart.png'), dpi=150)
            plt.close()
    
    def _create_technical_analysis_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create technical analysis sheet with indicators and conditional formatting."""
        ws = wb.create_sheet("Technical Analysis")
        ticker = data['ticker']
        
        # Title
        ws['A1'] = f"{ticker} - Technical Analysis"
        ws['A1'].style = self.styles['title']
        ws.merge_cells('A1:L1')
        
        # Get technical data
        tech_data = data.get('technical_data', pd.DataFrame())
        
        if tech_data.empty:
            ws['A3'] = "No technical data available"
            return
        
        # Prepare last 60 days of data
        display_data = tech_data.tail(60).copy()
        
        # Headers
        headers = ['Date', 'Close', 'Volume', 'SMA_20', 'SMA_50', 'SMA_200', 
                  'RSI', 'MACD', 'MACD_Signal', 'BB_Upper', 'BB_Lower', 'ATR']
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.style = self.styles['header']
        
        # Add data
        row = 4
        for idx, data_row in display_data.iterrows():
            ws.cell(row=row, column=1, value=idx.tz_localize(None).strftime('%Y-%m-%d') if hasattr(idx, 'tz_localize') else idx.strftime('%Y-%m-%d') if hasattr(idx, 'strftime') else str(idx))            
            ws.cell(row=row, column=2, value=float(data_row.get('Close', 0)))
            ws.cell(row=row, column=3, value=float(data_row.get('Volume', 0)))
            
            # Moving averages
            for col, indicator in enumerate(['SMA_20', 'SMA_50', 'SMA_200'], 4):
                value = data_row.get(indicator)
                if pd.notna(value):
                    ws.cell(row=row, column=col, value=float(value))
            
            # Technical indicators
            for col, indicator in enumerate(['RSI', 'MACD', 'MACD_Signal', 'BB_Upper', 'BB_Lower', 'ATR'], 7):
                value = data_row.get(indicator)
                if pd.notna(value):
                    ws.cell(row=row, column=col, value=float(value))
            
            row += 1
        
        # Apply conditional formatting
        last_row = row - 1
        
        # RSI formatting (column 7)
        # Oversold (< 30) - Green
        ws.conditional_formatting.add(f'G4:G{last_row}',
            CellIsRule(operator='lessThan', formula=['30'], 
                      fill=PatternFill(start_color='FF27AE60', end_color='FF27AE60', fill_type='solid')))
        
        # Overbought (> 70) - Red
        ws.conditional_formatting.add(f'G4:G{last_row}',
            CellIsRule(operator='greaterThan', formula=['70'], 
                      fill=PatternFill(start_color='FFE74C3C', end_color='FFE74C3C', fill_type='solid')))
        
        # Volume bars (column 3)
        ws.conditional_formatting.add(f'C4:C{last_row}',
            DataBarRule(start_type='min', end_type='max',
                       color="FF3498DB", showValue=True, minLength=None, maxLength=None))
        
        # Price color scale (column 2)
        ws.conditional_formatting.add(f'B4:B{last_row}',
            ColorScaleRule(start_type='min', start_color='FFE74C3C',
                          mid_type='percentile', mid_value=50, mid_color='FFF39C12',
                          end_type='max', end_color='FF27AE60'))
        
        # Format numbers
        for row in range(4, last_row + 1):
            # Price columns
            for col in [2, 4, 5, 6, 10, 11]:
                cell = ws.cell(row=row, column=col)
                cell.number_format = self.number_formats['currency']
            
            # Volume
            ws.cell(row=row, column=3).number_format = self.number_formats['integer']
            
            # Indicators
            for col in [7, 8, 9, 12]:
                cell = ws.cell(row=row, column=col)
                cell.number_format = self.number_formats['number']
        
        # Add signal interpretation section
        row = last_row + 3
        ws[f'A{row}'] = "Signal Interpretation"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        interpretations = self._generate_technical_interpretations(data)
        for interpretation in interpretations:
            ws[f'A{row}'] = interpretation
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            row += 1
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            # Find the first non-merged cell to get the column letter
            column_letter = None
            for cell in column_cells:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break
            if not column_letter:
                continue  # skip if all cells are merged

            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_performance_metrics_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create performance metrics sheet with enhanced risk metrics."""
        ws = wb.create_sheet("Performance Metrics")
        ticker = data['ticker']
        summary = data['summary_statistics']
        
        # Title
        ws['A1'] = f"{ticker} - Performance & Risk Metrics"
        ws['A1'].style = self.styles['title']
        ws.merge_cells('A1:H1')
        
        # Returns Analysis
        row = 3
        ws[f'A{row}'] = "Returns Analysis"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        # Headers
        headers = ['Period', 'Return (%)', 'Annualized Return (%)', 'Best Day', 'Worst Day', 'Positive Days %']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).style = self.styles['header']
        
        # Calculate additional metrics
        tech_data = data.get('technical_data', pd.DataFrame())
        if not tech_data.empty and 'Close' in tech_data.columns:
            # Ensure data is sorted by date
            tech_data = tech_data.sort_index()
            daily_returns = tech_data['Close'].pct_change().dropna()
            
            returns_data = []
            periods = [
                ('1 Day', 1), ('1 Week', 5), ('1 Month', 21), 
                ('3 Months', 63), ('6 Months', 126), ('1 Year', 252)
            ]
            
            for period_name, days in periods:
                if len(tech_data) >= days:
                    # Calculate period return properly
                    if days == 1:
                        # For 1 day, use the last day's return
                        period_return = daily_returns.iloc[-1] if len(daily_returns) > 0 else 0  # Already in decimal form
                    else:
                        # For longer periods, calculate cumulative return
                        start_price = tech_data['Close'].iloc[-days-1] if len(tech_data) > days else tech_data['Close'].iloc[0]
                        end_price = tech_data['Close'].iloc[-1]
                        period_return = ((end_price / start_price) - 1)  # Keep as decimal
                    
                    # Get returns for this period
                    period_daily_returns = daily_returns.tail(days)
                    
                    # Calculate annualized return properly
                    if days == 1:
                        # Don't annualize single day returns
                        annualized_return = period_return * 252  # Simple annualization for 1 day
                    else:
                        # Proper annualization formula
                        annualized_return = ((1 + period_return) ** (252/days)) - 1  # Keep as decimal
                    
                    # Calculate positive days percentage (should be 0-100)
                    positive_days_count = (period_daily_returns > 0).sum()
                    total_days = len(period_daily_returns)
                    positive_days_pct = (positive_days_count / total_days) if total_days > 0 else 0
                    
                    returns_data.append({
                        'Period': period_name,
                        'Return': period_return,
                        'Annualized': annualized_return,
                        'Best Day': period_daily_returns.max() if len(period_daily_returns) > 0 else 0,  # Already in decimal form
                        'Worst Day': period_daily_returns.min() if len(period_daily_returns) > 0 else 0,  # Already in decimal form
                        'Positive Days': positive_days_pct  # Already in decimal form (0.60 for 60%)
                    })
                else:
                    # Not enough data for this period
                    returns_data.append({
                        'Period': period_name,
                        'Return': 'N/A',
                        'Annualized': 'N/A',
                        'Best Day': 'N/A',
                        'Worst Day': 'N/A',
                        'Positive Days': 'N/A'
                    })
            
            # Add data to sheet
            row += 1
            for ret_data in returns_data:
                ws.cell(row=row, column=1, value=ret_data['Period'])
                
                # Handle N/A values
                for col, key in enumerate(['Return', 'Annualized', 'Best Day', 'Worst Day', 'Positive Days'], 2):
                    value = ret_data[key]
                    if value != 'N/A':
                        ws.cell(row=row, column=col, value=value)
                        # Format as percentage (Excel will multiply by 100 for display)
                        ws.cell(row=row, column=col).number_format = '0.00%'
                        
                        # Color coding for Return column
                        if col == 2 and isinstance(value, (int, float)):
                            if value > 0:
                                ws.cell(row=row, column=col).font = Font(color=self.colors['success'])
                            elif value < 0:
                                ws.cell(row=row, column=col).font = Font(color=self.colors['danger'])
                    else:
                        ws.cell(row=row, column=col, value='N/A')
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                
                row += 1
        
        # Risk Metrics
        row += 2
        ws[f'A{row}'] = "Risk Metrics"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        risk_headers = ['Metric', 'Value', 'Description', 'Interpretation']
        for col, header in enumerate(risk_headers, 1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).style = self.styles['header']
        
        # Calculate additional risk metrics
        volatility = summary.get('volatility_annual', 0)
        sharpe = summary.get('sharpe_ratio', 0)
        max_dd = summary.get('max_drawdown', 0)
        
        risk_metrics = [
            {
                'Metric': 'Annual Volatility',
                'Value': f"{volatility:.2f}%",
                'Description': 'Standard deviation of returns annualized',
                'Interpretation': self._interpret_volatility(volatility)
            },
            {
                'Metric': 'Sharpe Ratio',
                'Value': f"{sharpe:.2f}",
                'Description': 'Risk-adjusted return (assumes 2% risk-free rate)',
                'Interpretation': self._interpret_sharpe(sharpe)
            },
            {
                'Metric': 'Maximum Drawdown',
                'Value': f"{max_dd:.2f}%",
                'Description': 'Largest peak-to-trough decline',
                'Interpretation': self._interpret_drawdown(max_dd)
            },
            {
                'Metric': 'Beta',
                'Value': f"{summary.get('beta', 'N/A')}",
                'Description': 'Correlation with market (S&P 500)',
                'Interpretation': self._interpret_beta(summary.get('beta', 1))
            },
            {
                'Metric': 'Value at Risk (95%)',
                'Value': f"{summary.get('var_95', 'N/A')}%",
                'Description': '95% confidence worst daily loss',
                'Interpretation': 'Maximum expected loss in 95% of trading days'
            }
        ]
        
        row += 1
        for risk_data in risk_metrics:
            ws.cell(row=row, column=1, value=risk_data['Metric'])
            ws.cell(row=row, column=2, value=risk_data['Value'])
            ws.cell(row=row, column=3, value=risk_data['Description'])
            ws.cell(row=row, column=4, value=risk_data['Interpretation'])
            ws.merge_cells(f'D{row}:G{row}')
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
            row += 1
        
        # Comparison to Market
        row += 2
        ws[f'A{row}'] = "Market Comparison"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:H{row}')
        
        row += 1
        comparison_note = (
            f"Note: Risk metrics are compared against market benchmarks. "
            f"S&P 500 typical volatility: 15-20%, Sharpe Ratio: 0.5-1.0"
        )
        ws[f'A{row}'] = comparison_note
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws[f'A{row}'].font = Font(italic=True, size=10)
        
        # Add calculation notes
        row += 2
        ws[f'A{row}'] = "Calculation Notes:"
        ws[f'A{row}'].font = Font(bold=True, size=10)
        row += 1
        notes = [
            "• Returns are calculated based on closing prices",
            "• Positive Days % shows the percentage of days with positive returns in the period",
            "• Annualized returns use compound annual growth rate (CAGR) formula",
            "• Best/Worst Day shows the single day performance within each period"
        ]
        for note in notes:
            ws[f'A{row}'] = note
            ws[f'A{row}'].font = Font(size=9)
            ws.merge_cells(f'A{row}:H{row}')
            row += 1
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            # Find the first non-merged cell to get the column letter
            column_letter = None
            for cell in column_cells:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break
            if not column_letter:
                continue  # skip if all cells are merged

            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

    
    def _create_raw_data_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create raw data sheet with conditional formatting."""
        ws = wb.create_sheet("Raw Data")
        ticker = data['ticker']
        
        # Title
        ws['A1'] = f"{ticker} - Historical Price Data with Event Analysis"
        ws['A1'].style = self.styles['title']
        ws.merge_cells('A1:L1')
        
        # Get raw data
        raw_data = data.get('technical_data', data.get('raw_data', pd.DataFrame()))
        if raw_data.empty:
            ws['A3'] = "No raw data available"
            return
        
        display_data = raw_data.copy()
        
        # Calculate daily change percentage if not already present
        if 'Daily Change %' not in display_data.columns:
            display_data['Daily Change %'] = display_data['Close'].pct_change()
        
        # Headers
        headers = ['Date', 'Open', 'High', 'Low', 'Close', 'Volume', 'Daily Change %', 
                'Event Type', 'Sentiment', 'Confidence', 'Impact', 'News Count', 'Sentiment_Overall', 'Sentiment_Financial', 'Sentiment_Confidence', 
                'Sentiment_Relevance', 'Key_Phrases']
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
            ws.cell(row=3, column=col).style = self.styles['header']
        
        # Write data rows
        row = 4
        for idx, data_row in display_data.iterrows():
            # Date
            ws.cell(row=row, column=1, value=idx.tz_localize(None).strftime('%Y-%m-%d') if hasattr(idx, 'tz_localize') else idx.strftime('%Y-%m-%d') if hasattr(idx, 'strftime') else str(idx))
            
            # Price data
            ws.cell(row=row, column=2, value=float(data_row.get('Open', 0)))
            ws.cell(row=row, column=3, value=float(data_row.get('High', 0)))
            ws.cell(row=row, column=4, value=float(data_row.get('Low', 0)))
            ws.cell(row=row, column=5, value=float(data_row.get('Close', 0)))
            ws.cell(row=row, column=6, value=int(data_row.get('Volume', 0)))
            
            # Daily change percentage
            daily_change = data_row.get('Daily Change %', data_row.get('Daily_Change_%', 0))
            if pd.notna(daily_change):
                ws.cell(row=row, column=7, value=float(daily_change))
            else:
                ws.cell(row=row, column=7, value=0)
            
            # Event analysis columns (if available)
            if 'Event_Type' in display_data.columns:
                ws.cell(row=row, column=8, value=str(data_row.get('Event_Type', '')))
                
                # Sentiment with color coding
                sentiment = data_row.get('Sentiment', '')
                sentiment_cell = ws.cell(row=row, column=9, value=str(sentiment))
                if sentiment == 'Bullish':
                    sentiment_cell.font = Font(color=self.colors['success'])
                elif sentiment == 'Bearish':
                    sentiment_cell.font = Font(color=self.colors['danger'])
                
                # Confidence score
                confidence = data_row.get('Confidence_Score', 0)
                conf_cell = ws.cell(row=row, column=10, value=float(confidence))
                
                # Impact level
                impact = data_row.get('Impact_Level', '')
                impact_cell = ws.cell(row=row, column=11, value=str(impact))
                if impact == 'HIGH':
                    impact_cell.font = Font(color=self.colors['danger'], bold=True)
                elif impact == 'MEDIUM':
                    impact_cell.font = Font(color=self.colors['warning'])
                
                # News count
                ws.cell(row=row, column=12, value=int(data_row.get('News_Count', 0)))
                
                # Highlight significant events with news
                if data_row.get('Is_Significant', False) and data_row.get('News_Count', 0) > 0:
                    # Apply light background to entire row for significant events
                    fill_color = self.colors.get('event_bullish', 'FFC6EFCE') if sentiment == 'Bullish' else self.colors.get('event_bearish', 'FFFFC7CE')
                    for col in range(1, 13):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color=fill_color[2:],  # Remove 'FF' prefix
                            end_color=fill_color[2:],
                            fill_type="solid"
                        )
            else:
                # Fill empty event columns if not available
                for col in range(8, 13):
                    ws.cell(row=row, column=col, value='')
            
            row += 1

        last_row = row - 1

        # Number formatting
        for row in range(4, last_row + 1):
            # Price columns
            for col in range(2, 6):  # Open to Close
                ws.cell(row=row, column=col).number_format = self.number_formats['currency']
            
            # Volume
            ws.cell(row=row, column=6).number_format = self.number_formats['integer']
            
            # Daily change %
            ws.cell(row=row, column=7).number_format = '0.00%'
            
            # Confidence score (if event columns exist)
            if 'Event_Type' in display_data.columns:
                ws.cell(row=row, column=10).number_format = '0.00%'

        # Conditional formatting for daily change
        from openpyxl.formatting.rule import ColorScaleRule
        
        # Create a 3-color scale: red for negative, white for zero, green for positive
        color_scale_rule = ColorScaleRule(
            start_type='min',
            start_color='DC3545',  # Red
            mid_type='num',
            mid_value=0,
            mid_color='FFFFFF',    # White
            end_type='max',
            end_color='28A745'     # Green
        )
        
        ws.conditional_formatting.add(f'G4:G{last_row}', color_scale_rule)

        # Table
        table_ref = f"A3:L{last_row}"  # Updated to include all columns
        table = Table(displayName=f"{ticker}_RawData", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)

        # Auto-adjust columns with specific widths for certain columns
        for column_cells in ws.columns:
            column_letter = next((cell.column_letter for cell in column_cells if not isinstance(cell, MergedCell)), None)
            if not column_letter:
                continue
            
            # Calculate max length
            max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0)
            
            # Set specific widths for certain columns
            if column_letter == 'A':  # Date
                adjusted_width = 12
            elif column_letter == 'H':  # Event Type
                adjusted_width = 20
            elif column_letter == 'I':  # Sentiment
                adjusted_width = 12
            elif column_letter == 'J':  # Confidence
                adjusted_width = 12
            elif column_letter == 'K':  # Impact
                adjusted_width = 10
            elif column_letter == 'L':  # News Count
                adjusted_width = 12
            else:
                adjusted_width = min(max_length + 2, 30)
            
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add note about event analysis if available
        if 'Event_Type' in display_data.columns:
            row = last_row + 2
            ws[f'A{row}'] = f"Note: Highlighted rows indicate significant price movements (≥{self.significance_threshold}%) with news correlation"
            ws[f'A{row}'].font = Font(italic=True, size=9)
            ws.merge_cells(f'A{row}:L{row}')

    
    def _create_data_quality_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create data quality report sheet."""
        ws = wb.create_sheet("Data Quality")
        ticker = data['ticker']
        
        # Title
        ws['A1'] = f"{ticker} - Data Quality Report"
        ws['A1'].style = self.styles['title']
        ws.merge_cells('A1:F1')
        
        # Get data for quality analysis
        tech_data = data.get('technical_data', pd.DataFrame())
        
        if tech_data.empty:
            ws['A3'] = "No data available for quality analysis"
            return
        
        # Data Overview
        row = 3
        ws[f'A{row}'] = "Data Overview"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        overview_data = [
            ['Total Records', len(tech_data)],
            ['Date Range', f"{tech_data.index.min()} to {tech_data.index.max()}"],
            ['Trading Days', len(tech_data)],
            ['Expected Trading Days', self._calculate_expected_trading_days(tech_data.index.min(), tech_data.index.max())],
            ['Data Completeness', f"{(len(tech_data) / self._calculate_expected_trading_days(tech_data.index.min(), tech_data.index.max()) * 100):.1f}%"]
        ]
        
        for label, value in overview_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = str(value)
            ws.merge_cells(f'B{row}:D{row}')
            row += 1
        
        # Missing Data Analysis
        row += 1
        ws[f'A{row}'] = "Missing Data Analysis"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        # Headers
        headers = ['Column', 'Missing Count', 'Missing %', 'First Valid', 'Last Valid']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).style = self.styles['header']
        
        row += 1
        # Analyze each column
        columns_to_check = ['Open', 'High', 'Low', 'Close', 'Volume', 'SMA_20', 'SMA_50', 'SMA_200', 'RSI', 'MACD']
        for col_name in columns_to_check:
            if col_name in tech_data.columns:
                missing_count = tech_data[col_name].isna().sum()
                missing_pct = (missing_count / len(tech_data)) * 100
                first_valid = tech_data[col_name].first_valid_index()
                last_valid = tech_data[col_name].last_valid_index()
                
                ws.cell(row=row, column=1, value=col_name)
                ws.cell(row=row, column=2, value=missing_count)
                ws.cell(row=row, column=3, value=f"{missing_pct:.1f}%")
                ws.cell(row=row, column=4, value=str(first_valid) if first_valid else 'N/A')
                ws.cell(row=row, column=5, value=str(last_valid) if last_valid else 'N/A')
                
                # Color code based on missing percentage
                if missing_pct > 20:
                    ws.cell(row=row, column=3).font = Font(color=self.colors['danger'])
                elif missing_pct > 10:
                    ws.cell(row=row, column=3).font = Font(color=self.colors['warning'])
                else:
                    ws.cell(row=row, column=3).font = Font(color=self.colors['success'])
                
                row += 1
        
        # Data Integrity Checks
        row += 2
        ws[f'A{row}'] = "Data Integrity Checks"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        integrity_checks = []
        
        # Check for negative prices
        price_cols = ['Open', 'High', 'Low', 'Close']
        for col in price_cols:
            if col in tech_data.columns:
                negative_count = (tech_data[col] < 0).sum()
                integrity_checks.append({
                    'Check': f'Negative {col} prices',
                    'Result': 'PASS' if negative_count == 0 else 'FAIL',
                    'Details': f'{negative_count} negative values found'
                })
        
        # Check High >= Low
        if 'High' in tech_data.columns and 'Low' in tech_data.columns:
            invalid_hl = (tech_data['High'] < tech_data['Low']).sum()
            integrity_checks.append({
                'Check': 'High >= Low validation',
                'Result': 'PASS' if invalid_hl == 0 else 'FAIL',
                'Details': f'{invalid_hl} days where High < Low'
            })
        
        # Check for extreme price movements (>20% in a day)
        if 'Close' in tech_data.columns:
            daily_returns = tech_data['Close'].pct_change()
            extreme_moves = (daily_returns.abs() > 0.2).sum()
            integrity_checks.append({
                'Check': 'Extreme price movements (>20%)',
                'Result': 'WARNING' if extreme_moves > 0 else 'PASS',
                'Details': f'{extreme_moves} days with >20% price change'
            })
        
        # Check for zero volume days
        if 'Volume' in tech_data.columns:
            zero_volume = (tech_data['Volume'] == 0).sum()
            integrity_checks.append({
                'Check': 'Zero volume days',
                'Result': 'WARNING' if zero_volume > 0 else 'PASS',
                'Details': f'{zero_volume} days with zero volume'
            })
        
        # Add integrity check results
        for check in integrity_checks:
            ws[f'A{row}'] = check['Check']
            ws[f'C{row}'] = check['Result']
            ws[f'D{row}'] = check['Details']
            ws.merge_cells(f'D{row}:F{row}')
            
            # Color code results
            if check['Result'] == 'PASS':
                ws[f'C{row}'].font = Font(color=self.colors['success'], bold=True)
            elif check['Result'] == 'WARNING':
                ws[f'C{row}'].font = Font(color=self.colors['warning'], bold=True)
            else:
                ws[f'C{row}'].font = Font(color=self.colors['danger'], bold=True)
            
            row += 1
        
        # Summary
        row += 2
        ws[f'A{row}'] = "Data Quality Summary"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        summary_text = self._generate_data_quality_summary(tech_data, integrity_checks)
        ws[f'A{row}'] = summary_text
        ws.merge_cells(f'A{row}:F{row+3}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            # Find the first non-merged cell to get the column letter
            column_letter = None
            for cell in column_cells:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break
            if not column_letter:
                continue  # skip if all cells are merged

            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_metadata_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create metadata sheet with analysis information."""
        ws = wb.create_sheet("Metadata")
        ticker = data['ticker']
        
        # Title
        ws['A1'] = f"{ticker} - Analysis Metadata"
        ws['A1'].style = self.styles['title']
        ws.merge_cells('A1:D1')
        
        # Analysis Information
        row = 3
        ws[f'A{row}'] = "Analysis Information"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        metadata = [
            ['Report Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Analysis Tool', 'Stock Analyzer v1.0'],
            ['Data Source', 'Yahoo Finance'],
            ['Analysis Period', '5 Years'],
            ['Ticker Symbol', ticker],
            ['Company Name', data['company_info'].get('longName', 'N/A')],
            ['Report Type', 'Comprehensive Stock Analysis'],
            ['File Format', 'Excel (.xlsx)']
        ]
        
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:D{row}')
            row += 1
        
        # Technical Indicators Used
        row += 1
        ws[f'A{row}'] = "Technical Indicators Included"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        indicators = [
            'Simple Moving Averages (20, 50, 200 days)',
            'Relative Strength Index (RSI) - 14 day',
            'MACD (12, 26, 9)',
            'Bollinger Bands (20 day, 2 std dev)',
            'Average True Range (ATR) - 14 day',
            'Stochastic Oscillator (14, 3, 3)',
            'On Balance Volume (OBV)',
            'Commodity Channel Index (CCI)'
        ]
        
        for indicator in indicators:
            ws[f'A{row}'] = f"• {indicator}"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
        
        # Risk Metrics Calculated
        row += 1
        ws[f'A{row}'] = "Risk Metrics Calculated"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        risk_metrics = [
            'Annual Volatility (252-day)',
            'Sharpe Ratio (Risk-free rate: 2%)',
            'Maximum Drawdown',
            'Beta (vs S&P 500)',
            'Value at Risk (95% confidence)',
            'Daily, Weekly, Monthly Returns',
            'Risk-Adjusted Returns'
        ]
        
        for metric in risk_metrics:
            ws[f'A{row}'] = f"• {metric}"
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
        
        # Disclaimer
        row += 2
        ws[f'A{row}'] = "Disclaimer"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        disclaimer = (
            "This report is for informational purposes only and should not be considered as financial advice. "
            "Past performance is not indicative of future results. All investments carry risk, including the "
            "potential loss of principal. Please consult with a qualified financial advisor before making any "
            "investment decisions. The data and analysis provided in this report are based on historical "
            "information and may not reflect current market conditions."
        )
        ws[f'A{row}'] = disclaimer
        ws.merge_cells(f'A{row}:D{row+5}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws[f'A{row}'].font = Font(italic=True, size=10)
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            # Find the first non-merged cell to get the column letter
            column_letter = None
            for cell in column_cells:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break
            if not column_letter:
                continue  # skip if all cells are merged

            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Helper methods
    def _generate_technical_interpretations(self, data: Dict[str, Any]) -> List[str]:
        """Generate interpretations for technical indicators."""
        interpretations = []
        summary = data['summary_statistics']
        indicators = summary.get('technical_indicators', {})
        
        # RSI interpretation
        rsi = indicators.get('rsi')
        if rsi:
            if rsi > 70:
                interpretations.append(f"• RSI at {rsi:.2f} indicates overbought conditions - potential for price pullback")
            elif rsi < 30:
                interpretations.append(f"• RSI at {rsi:.2f} indicates oversold conditions - potential for price bounce")
            else:
                interpretations.append(f"• RSI at {rsi:.2f} indicates neutral momentum")
        
        # Moving average interpretation
        current_price = summary['current_price']
        sma_20 = indicators.get('sma_20')
        sma_50 = indicators.get('sma_50')
        sma_200 = indicators.get('sma_200')
        
        if sma_20 and sma_50 and sma_200:
            if current_price > sma_20 > sma_50 > sma_200:
                interpretations.append("• Strong uptrend: Price above all major moving averages in proper order")
            elif current_price < sma_20 < sma_50 < sma_200:
                interpretations.append("• Strong downtrend: Price below all major moving averages")
            else:
                interpretations.append("• Mixed trend signals from moving averages")
        
        # MACD interpretation
        macd = indicators.get('macd')
        macd_signal = indicators.get('macd_signal')
        if macd is not None and macd_signal is not None:
            if macd > macd_signal:
                interpretations.append("• MACD above signal line - bullish momentum")
            else:
                interpretations.append("• MACD below signal line - bearish momentum")
        
        return interpretations
    
    def _interpret_volatility(self, volatility: float) -> str:
        """Interpret volatility level."""
        if volatility < 10:
            return "Very low volatility - stable price movement"
        elif volatility < 20:
            return "Low to moderate volatility - normal for large cap stocks"
        elif volatility < 30:
            return "Moderate volatility - typical for growth stocks"
        elif volatility < 50:
            return "High volatility - significant price swings"
        else:
            return "Very high volatility - extreme price movements"
    
    def _interpret_sharpe(self, sharpe: float) -> str:
        """Interpret Sharpe ratio."""
        if sharpe < 0:
            return "Negative risk-adjusted returns"
        elif sharpe < 0.5:
            return "Poor risk-adjusted returns"
        elif sharpe < 1.0:
            return "Below average risk-adjusted returns"
        elif sharpe < 1.5:
            return "Good risk-adjusted returns"
        elif sharpe < 2.0:
            return "Very good risk-adjusted returns"
        else:
            return "Excellent risk-adjusted returns"
    
    def _interpret_drawdown(self, drawdown: float) -> str:
        """Interpret maximum drawdown."""
        drawdown = abs(drawdown)
        if drawdown < 10:
            return "Low maximum loss from peak"
        elif drawdown < 20:
            return "Moderate maximum loss - typical for stocks"
        elif drawdown < 30:
            return "Significant drawdown - higher risk"
        elif drawdown < 50:
            return "Large drawdown - very high risk"
        else:
            return "Extreme drawdown - extremely high risk"
    
    def _interpret_beta(self, beta: float) -> str:
        """Interpret beta value."""
        if isinstance(beta, str) or beta is None:
            return "Beta not available"
        
        if beta < 0.5:
            return "Low correlation with market - defensive stock"
        elif beta < 0.8:
            return "Below market volatility"
        elif beta < 1.2:
            return "Moves with the market"
        elif beta < 1.5:
            return "More volatile than market"
        else:
            return "Highly volatile compared to market"
    
    def _calculate_expected_trading_days(self, start_date, end_date) -> int:
        """Calculate expected number of trading days between dates."""
        # Rough estimate: 252 trading days per year
        days_diff = (end_date - start_date).days
        years = days_diff / 365.25
        return int(years * 252)
    
    def _generate_data_quality_summary(self, data: pd.DataFrame, integrity_checks: List[Dict]) -> str:
        """Generate summary text for data quality."""
        failed_checks = sum(1 for check in integrity_checks if check['Result'] == 'FAIL')
        warning_checks = sum(1 for check in integrity_checks if check['Result'] == 'WARNING')
        
        if failed_checks == 0 and warning_checks == 0:
            quality_status = "EXCELLENT - All data quality checks passed"
        elif failed_checks == 0:
            quality_status = f"GOOD - {warning_checks} warnings but no critical issues"
        else:
            quality_status = f"NEEDS ATTENTION - {failed_checks} failed checks, {warning_checks} warnings"
        
        missing_pct = (data.isna().sum().sum() / (len(data) * len(data.columns))) * 100
        
        summary = (
            f"Data Quality Status: {quality_status}\n\n"
            f"The dataset contains {len(data):,} trading days of data with {missing_pct:.1f}% "
            f"missing values overall. Price data (OHLCV) is complete, while technical indicators "
            f"have expected missing values at the beginning due to calculation requirements. "
            f"No critical data integrity issues were found that would impact the analysis."
        )
        
        return summary
    
    def _analyze_price_events(self, tech_data: pd.DataFrame, ticker: str) -> pd.DataFrame:
        """Analyze significant price movements with enhanced sentiment and intelligent analysis."""
        try:
            self.logger.info(f"Analyzing price events for {ticker} using enhanced system")
            
            # Create a copy to avoid modifying original data
            analyzed_data = tech_data.copy()
            
            # Ensure we have required columns
            if 'Close' not in analyzed_data.columns or 'Open' not in analyzed_data.columns:
                self.logger.warning("Missing required price columns for event analysis")
                return tech_data
            
            # Calculate daily price change percentage
            analyzed_data['Daily_Change_%'] = (
                (analyzed_data['Close'] - analyzed_data['Open']) / analyzed_data['Open'] * 100
            ).round(2)
            
            # Handle any infinite or NaN values
            analyzed_data['Daily_Change_%'] = analyzed_data['Daily_Change_%'].replace(
                [float('inf'), float('-inf')], 0
            ).fillna(0)
            
            # Use dynamic threshold from enhanced analyzer
            current_threshold = self.enhanced_analyzer.threshold_manager.get_current_threshold()
            
            # Identify significant movements using dynamic threshold
            analyzed_data['Is_Significant'] = (
                abs(analyzed_data['Daily_Change_%']) >= current_threshold
            )
            
            # Initialize enhanced event analysis columns
            analyzed_data['Event_Type'] = ''
            analyzed_data['Event_Reason'] = ''
            analyzed_data['Sentiment'] = ''
            analyzed_data['Confidence_Score'] = 0.0
            analyzed_data['Impact_Level'] = ''
            analyzed_data['News_Count'] = 0
            
            # NEW: Add sentiment analysis columns
            analyzed_data['Sentiment_Overall'] = ''
            analyzed_data['Sentiment_Financial'] = ''
            analyzed_data['Sentiment_Confidence'] = 0.0
            analyzed_data['Sentiment_Relevance'] = 0.0
            analyzed_data['Key_Phrases'] = ''
            analyzed_data['Analysis_Method'] = ''  # GPT Knowledge/Full Scrape/Fallback
            analyzed_data['Analysis_Phase'] = ''   # Learning/Knowledge
            
            # Get significant events (most recent first for performance)
            significant_events = analyzed_data[analyzed_data['Is_Significant']].copy()
            significant_events = significant_events.sort_index(ascending=False).head(self.max_days_to_analyze)
            
            phase_info = self.enhanced_analyzer.threshold_manager.get_phase_info()
            self.logger.info(
                f"Found {len(significant_events)} significant events to analyze "
                f"(Threshold: {current_threshold}%, Phase: {phase_info['current_phase']})"
            )
            
            # Analyze each significant event using enhanced system
            events_analyzed = 0
            high_confidence_events = 0
            gpt_sufficient_events = 0
            
            for idx, row in significant_events.iterrows():
                try:
                    # Create PriceEvent object
                    event = PriceEvent(
                        date=row.name if isinstance(row.name, datetime) else pd.to_datetime(row.name),
                        ticker=ticker,
                        open_price=float(row['Open']),
                        close_price=float(row['Close']),
                        change_percent=float(row['Daily_Change_%']),
                        volume=int(row.get('Volume', 0)),
                        is_significant=True
                    )
                    
                    self.logger.info(
                        f"Processing {ticker} event on {event.date.date()} "
                        f"({event.change_percent:+.2f}%) - {phase_info['current_phase']} phase"
                    )
                    
                    # Use enhanced intelligent analysis
                    analysis = self.enhanced_analyzer.analyze_event_intelligent(event)
                    
                    if analysis:
                        # Update the dataframe with enhanced analysis results
                        analyzed_data.loc[idx, 'Event_Type'] = analysis.event_type
                        analyzed_data.loc[idx, 'Event_Reason'] = analysis.event_reason
                        analyzed_data.loc[idx, 'Sentiment'] = analysis.sentiment
                        analyzed_data.loc[idx, 'Confidence_Score'] = analysis.confidence_score
                        analyzed_data.loc[idx, 'Impact_Level'] = analysis.impact_level
                        analyzed_data.loc[idx, 'News_Count'] = analysis.news_sources_count
                        analyzed_data.loc[idx, 'Analysis_Method'] = analysis.analysis_method
                        analyzed_data.loc[idx, 'Analysis_Phase'] = phase_info['current_phase']
                        
                        # NEW: Get sentiment data if available (for Full Scrape method)
                        if analysis.analysis_method == "full_scrape":
                            # Get the articles with sentiment from the news bridge
                            try:
                                articles, sentiment_summary = news_bridge.get_news_with_sentiment(
                                    ticker=ticker,
                                    event_date=event.date,
                                    lookback_days=3
                                )
                                
                                # Check for errors in the response
                                if sentiment_summary.get('error'):
                                    self.logger.warning(
                                        f"News analysis error for {ticker} on {event.date}: "
                                        f"{sentiment_summary['error']}"
                                    )
                                elif articles and sentiment_summary:
                                    # Get overall sentiment metrics
                                    overall_assessment = sentiment_summary.get('overall_assessment', {})
                                    financial_sentiment = sentiment_summary.get('financial_sentiment_breakdown', {})
                                    quality_metrics = sentiment_summary.get('quality_metrics', {})
                                    
                                    analyzed_data.loc[idx, 'Sentiment_Overall'] = overall_assessment.get('market_sentiment', '')
                                    analyzed_data.loc[idx, 'Sentiment_Financial'] = overall_assessment.get('market_sentiment', '')
                                    analyzed_data.loc[idx, 'Sentiment_Confidence'] = sentiment_summary.get('confidence', 0)
                                    
                                    # Get average sentiment relevance from articles
                                    if articles:
                                        avg_relevance = sum(a.get('sentiment_relevance', 0) for a in articles) / len(articles)
                                        analyzed_data.loc[idx, 'Sentiment_Relevance'] = avg_relevance
                                        
                                        # Collect key phrases from top articles
                                        all_phrases = []

                                        # First try to get from overall sentiment summary
                                        if sentiment_summary and 'key_themes' in sentiment_summary:
                                            all_phrases.extend(sentiment_summary.get('key_themes', [])[:3])

                                        # Then get from individual articles
                                        for article in articles[:3]:  # Top 3 articles
                                            # Check different possible locations for key phrases
                                            if 'key_phrases' in article:
                                                all_phrases.extend(article['key_phrases'][:2])
                                            elif 'sentiment' in article and isinstance(article['sentiment'], dict):
                                                if 'key_phrases' in article['sentiment']:
                                                    all_phrases.extend(article['sentiment']['key_phrases'][:2])
                                            
                                            # Also check for tags or categories as key phrases
                                            if 'tags' in article:
                                                all_phrases.extend(article.get('tags', [])[:1])
                                            if 'categories' in article:
                                                all_phrases.extend(article.get('categories', [])[:1])

                                        # Remove duplicates and limit to 5
                                        unique_phrases = []
                                        for phrase in all_phrases:
                                            if phrase and phrase not in unique_phrases:
                                                unique_phrases.append(phrase)

                                        analyzed_data.loc[idx, 'Key_Phrases'] = ', '.join(unique_phrases[:5]) if unique_phrases else 'Market sentiment, Price movement'
                            except Exception as e:
                                self.logger.warning(f"Could not get sentiment details for {idx}: {e}")
                        
                        events_analyzed += 1
                        if analysis.confidence_score >= 0.7:
                            high_confidence_events += 1
                        if analysis.analysis_method == "gpt_knowledge":
                            gpt_sufficient_events += 1
                            
                    else:
                        # Use fallback analysis if enhanced analysis fails
                        fallback = self.enhanced_analyzer.create_fallback_analysis(event)
                        analyzed_data.loc[idx, 'Event_Type'] = fallback.event_type
                        analyzed_data.loc[idx, 'Event_Reason'] = fallback.event_reason
                        analyzed_data.loc[idx, 'Sentiment'] = fallback.sentiment
                        analyzed_data.loc[idx, 'Confidence_Score'] = fallback.confidence_score
                        analyzed_data.loc[idx, 'Impact_Level'] = fallback.impact_level
                        analyzed_data.loc[idx, 'News_Count'] = 0
                        analyzed_data.loc[idx, 'Analysis_Method'] = fallback.analysis_method
                        analyzed_data.loc[idx, 'Analysis_Phase'] = phase_info['current_phase']
                        
                        events_analyzed += 1
                        
                except Exception as e:
                    self.logger.error(f"Error analyzing event at {idx}: {e}")
                    continue
            
            # Get enhanced statistics
            enhanced_stats = self.enhanced_analyzer.get_enhanced_stats()
            efficiency_metrics = enhanced_stats.get('efficiency_metrics', {})
            
            self.logger.info(
                f"Enhanced analysis complete: {events_analyzed} events analyzed, "
                f"{high_confidence_events} with high confidence, "
                f"{gpt_sufficient_events} using GPT knowledge only, "
                f"GPT success rate: {efficiency_metrics.get('gpt_success_rate', 0):.1f}%"
            )
            
            return analyzed_data
            
        except Exception as e:
            self.logger.error(f"Error in enhanced price event analysis: {e}")
            return tech_data

    def _create_sentiment_analysis_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create sentiment analysis sheet with article-by-article breakdown."""
        ws = wb.create_sheet("Sentiment Analysis")
        ticker = data['ticker']
        
        # Title
        ws['A1'] = f"{ticker} - Sentiment Analysis Report"
        ws['A1'].style = self.styles['title']
        ws.merge_cells('A1:M1')  # Extended to M to accommodate new layout
        
        # Add generation date and description
        ws['A2'] = f"AI-powered sentiment analysis with GPT intelligence and dynamic thresholds"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:M2')
        
        # Get event analysis data with sentiment information
        event_data = data.get('event_analysis', pd.DataFrame())
        
        if event_data.empty:
            ws['A4'] = "No sentiment data available - run event analysis first"
            ws['A4'].font = Font(italic=True)
            return
        
        # Filter for events with sentiment data
        sentiment_events = event_data[
            (event_data.get('Sentiment_Overall', '') != '') | 
            (event_data.get('Analysis_Method', '') != '')
        ].copy()
        
        if sentiment_events.empty:
            ws['A4'] = "No sentiment analysis data found in events"
            ws['A4'].font = Font(italic=True)
            return
        
        # Overall Sentiment Summary Section
        row = 4
        ws[f'A{row}'] = "Overall Sentiment Summary"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:M{row}')
        
        row += 1
        # Calculate sentiment statistics from event data
        total_events = len(sentiment_events)
        
        if total_events > 0:
            # Count different sentiment types
            bullish_count = len(sentiment_events[sentiment_events['Sentiment'] == 'Bullish'])
            bearish_count = len(sentiment_events[sentiment_events['Sentiment'] == 'Bearish'])
            neutral_count = total_events - bullish_count - bearish_count
            
            # High confidence events (from event analysis)
            high_confidence_count = len(sentiment_events[sentiment_events['Confidence_Score'] >= 0.7])
            
            # Average confidence from event analysis
            avg_confidence = sentiment_events['Confidence_Score'].mean()
            
            # News coverage statistics
            events_with_news = len(sentiment_events[sentiment_events['News_Count'] > 0])
            total_news_articles = sentiment_events['News_Count'].sum()
            
            # Sentiment confidence average (from detailed sentiment analysis)
            sentiment_confidence_scores = sentiment_events['Sentiment_Confidence'].fillna(0)
            avg_sentiment_confidence = sentiment_confidence_scores.mean()
            
            summary_stats = [
                ['Total Analyzed Events:', total_events],
                ['Bullish Events:', f"{bullish_count} ({bullish_count/total_events:.1%})"],
                ['Bearish Events:', f"{bearish_count} ({bearish_count/total_events:.1%})"],
                ['Neutral Events:', f"{neutral_count} ({neutral_count/total_events:.1%})"],
                ['High Confidence Events (≥70%):', f"{high_confidence_count} ({high_confidence_count/total_events:.1%})"],
                ['Average Analysis Confidence:', f"{avg_confidence:.2%}"],
                ['Average Sentiment Confidence:', f"{avg_sentiment_confidence:.2%}"],
                ['Events with News Coverage:', f"{events_with_news} ({events_with_news/total_events:.1%})"],
                ['Total News Articles Analyzed:', int(total_news_articles)]
            ]
        else:
            summary_stats = [['No sentiment events available for analysis', '']]
        
        # Display summary statistics
        for label, value in summary_stats:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:D{row}')
            row += 1
        
        # Sentiment Distribution Chart Data
        if total_events > 0:
            ws[f'G{row-9}'] = "Sentiment Distribution"
            ws[f'G{row-9}'].style = self.styles['subheader']
            ws.merge_cells(f'G{row-9}:M{row-9}')
            
            dist_row = row - 8
            ws[f'G{dist_row}'] = "Bullish"
            ws[f'H{dist_row}'] = bullish_count
            ws[f'I{dist_row}'] = f"{bullish_count/total_events:.1%}"
            ws[f'H{dist_row}'].font = Font(color=self.colors['success'], bold=True)
            
            dist_row += 1
            ws[f'G{dist_row}'] = "Bearish"
            ws[f'H{dist_row}'] = bearish_count
            ws[f'I{dist_row}'] = f"{bearish_count/total_events:.1%}"
            ws[f'H{dist_row}'].font = Font(color=self.colors['danger'], bold=True)
            
            dist_row += 1
            ws[f'G{dist_row}'] = "Neutral"
            ws[f'H{dist_row}'] = neutral_count
            ws[f'I{dist_row}'] = f"{neutral_count/total_events:.1%}"
            ws[f'H{dist_row}'].font = Font(color=self.colors['warning'])
        
        # Intelligence Metrics Section
        row += 2
        ws[f'A{row}'] = "Intelligence Metrics"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:M{row}')
        
        row += 1
        # Get intelligence stats from enhanced_analyzer
        enhanced_stats = self.enhanced_analyzer.get_enhanced_stats()
        threshold_info = enhanced_stats.get('threshold_manager', {})
        efficiency_metrics = enhanced_stats.get('efficiency_metrics', {})
        
        # Count analysis methods from our event data
        gpt_knowledge_events = len(sentiment_events[sentiment_events['Analysis_Method'] == 'gpt_knowledge'])
        full_scrape_events = len(sentiment_events[sentiment_events['Analysis_Method'] == 'full_scrape'])
        fallback_events = len(sentiment_events[sentiment_events['Analysis_Method'] == 'fallback'])
        learning_phase_events = len(sentiment_events[sentiment_events['Analysis_Phase'] == 'learning'])
        knowledge_phase_events = len(sentiment_events[sentiment_events['Analysis_Phase'] == 'knowledge'])
        
        intel_metrics = [
            ['Current Analysis Phase:', threshold_info.get('current_phase', 'Unknown').title()],
            ['Current Threshold:', f"{threshold_info.get('current_threshold', 0):.1f}%"],
            ['Learning Phase Events:', learning_phase_events],
            ['Knowledge Phase Events:', knowledge_phase_events],
            ['GPT Knowledge Sufficient:', gpt_knowledge_events],
            ['Full Scraping Used:', full_scrape_events],
            ['Fallback Analysis Used:', fallback_events],
            ['GPT Success Rate:', f"{efficiency_metrics.get('gpt_success_rate', 0):.1f}%"],
            ['Estimated Cost Savings:', f"${efficiency_metrics.get('estimated_cost_savings', 0):.2f}"],
        ]
        
        for label, value in intel_metrics:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:D{row}')
            row += 1
        
        # Event-by-Event Sentiment Breakdown
        row += 2
        ws[f'A{row}'] = "Event-by-Event Sentiment Breakdown"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:M{row}')
        
        row += 1
        header_row = row
        
        # Headers for event table
        headers = ['Date', 'Price %', 'Event Type', 'Sentiment', 'Conf.', 
                'News', 'Method', 'Phase', 'Impact', 
                'Sent.Score', 'Key Phrases', 'Event Reason Summary', 'Full Analysis']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.style = self.styles['header']
        
        # Sort events by date descending (most recent first)
        sentiment_events = sentiment_events.sort_index(ascending=False)
        
        # Add event data
        row += 1
        event_data_start_row = row
        
        for idx, event in sentiment_events.iterrows():
            # Date
            event_date = idx.strftime('%Y-%m-%d') if hasattr(idx, 'strftime') else str(idx)
            ws.cell(row=row, column=1, value=event_date)
            
            # Price Change % (shortened header)
            price_change = event.get('Daily_Change_%', 0)
            change_cell = ws.cell(row=row, column=2, value=f"{price_change:+.1f}%")
            if price_change > 0:
                change_cell.font = Font(color=self.colors['success'])
            else:
                change_cell.font = Font(color=self.colors['danger'])
            
            # Event Type
            event_type = event.get('Event_Type', 'Unknown')
            # Truncate long event types
            if len(event_type) > 20:
                event_type = event_type[:17] + "..."
            ws.cell(row=row, column=3, value=event_type)
            
            # Sentiment
            sentiment = event.get('Sentiment', 'Neutral')
            sentiment_cell = ws.cell(row=row, column=4, value=sentiment)
            if sentiment == 'Bullish':
                sentiment_cell.font = Font(color=self.colors['success'], bold=True)
            elif sentiment == 'Bearish':
                sentiment_cell.font = Font(color=self.colors['danger'], bold=True)
            else:
                sentiment_cell.font = Font(color=self.colors['warning'])
            
            # Confidence (shortened)
            confidence = event.get('Confidence_Score', 0)
            conf_cell = ws.cell(row=row, column=5, value=f"{confidence:.0%}")
            if confidence >= 0.7:
                conf_cell.font = Font(color=self.colors['success'], bold=True)
            elif confidence >= 0.5:
                conf_cell.font = Font(color=self.colors['warning'])
            else:
                conf_cell.font = Font(color=self.colors['danger'])
            
            # News Count (shortened)
            news_count = event.get('News_Count', 0)
            ws.cell(row=row, column=6, value=int(news_count))
            
            # Analysis Method (shortened)
            analysis_method = event.get('Analysis_Method', 'Unknown')
            method_cell = ws.cell(row=row, column=7)
            if analysis_method == 'gpt_knowledge':
                method_cell.font = Font(color=self.colors['success'])
                method_cell.value = 'GPT'
            elif analysis_method == 'full_scrape':
                method_cell.font = Font(color=self.colors['warning'])
                method_cell.value = 'Scrape'
            else:
                method_cell.font = Font(color=self.colors['danger'])
                method_cell.value = 'Fallback'
            
            # Analysis Phase
            phase = event.get('Analysis_Phase', 'Unknown')
            phase_cell = ws.cell(row=row, column=8, value=phase[:5].title())  # Shortened
            if phase == 'learning':
                phase_cell.font = Font(color=self.colors['secondary'])
            else:
                phase_cell.font = Font(color=self.colors['primary'])
            
            # Impact Level
            impact = event.get('Impact_Level', 'LOW')
            impact_cell = ws.cell(row=row, column=9, value=impact[:3])  # Shortened
            if impact == 'HIGH':
                impact_cell.font = Font(color=self.colors['danger'], bold=True)
            elif impact == 'MEDIUM':
                impact_cell.font = Font(color=self.colors['warning'])
            
            # Sentiment Score (shortened header)
            sentiment_confidence = event.get('Sentiment_Confidence', 0)
            if sentiment_confidence > 0:
                ws.cell(row=row, column=10, value=f"{sentiment_confidence:.0%}")
            else:
                ws.cell(row=row, column=10, value='N/A')
            
            # Key Phrases
            key_phrases = event.get('Key_Phrases', '')
            if not key_phrases or key_phrases == '':
                event_type_full = event.get('Event_Type', '')
                if event_type_full and event_type_full != 'Unknown':
                    key_phrases = f"{event_type_full}, {sentiment}"
                else:
                    key_phrases = 'Market movement'
            # Limit key phrases
            phrases_text = key_phrases[:60] + "..." if len(key_phrases) > 60 else key_phrases
            phrases_cell = ws.cell(row=row, column=11, value=phrases_text)
            phrases_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Event Reason Summary (truncated for main view)
            event_reason = event.get('Event_Reason', 'No analysis available')
            # Show first 150 characters in summary column
            reason_summary = event_reason[:150] + "..." if len(event_reason) > 150 else event_reason
            reason_cell = ws.cell(row=row, column=12, value=reason_summary)
            reason_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            
            # Full Analysis (complete text in separate column)
            full_reason_cell = ws.cell(row=row, column=13, value=event_reason)
            full_reason_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            
            # Set reasonable row height
            ws.row_dimensions[row].height = 40  # Good height for readability
            
            # Highlight high-confidence, high-impact events
            if confidence >= 0.7 and impact == 'HIGH':
                for col in range(1, 14):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color=self.colors['light'][2:],
                        end_color=self.colors['light'][2:], 
                        fill_type="solid"
                    )
            
            row += 1
        
        # Add AutoFilter to the event table for better navigation
        ws.auto_filter.ref = f"A{header_row}:M{row-1}"
        
        # No freeze panes - natural scrolling like a normal page
        
        # Add analysis notes
        row += 2
        ws[f'A{row}'] = "Analysis Notes"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:M{row}')
        
        row += 1
        notes = [
            "• GPT: Fast analysis using AI training data (cost-effective)",
            "• Scrape: Comprehensive analysis with fresh news articles (detailed but costly)",
            "• Fallback: Basic analysis when other methods fail",
            "• Learning Phase: Lower threshold (3%) to gather training data",
            "• Knowledge Phase: Higher threshold (7.5%) using accumulated intelligence",
            "• Confidence Score: AI confidence in the event analysis accuracy",
            "• Sent.Score: Separate confidence in sentiment analysis from news articles",
            "• Event Reason Summary: First 150 characters of analysis. See 'Full Analysis' column for complete text",
            f"• Current threshold: {threshold_info.get('current_threshold', 0):.1f}% in {threshold_info.get('current_phase', 'unknown')} phase"
        ]
        
        for note in notes:
            ws[f'A{row}'] = note
            ws[f'A{row}'].font = Font(size=9, italic=True)
            ws.merge_cells(f'A{row}:M{row}')
            row += 1
        
        # Column widths - full width for better readability
        column_widths = {
            'A': 12,   # Date
            'B': 10,   # Price %
            'C': 20,   # Event Type
            'D': 12,   # Sentiment
            'E': 10,   # Conf.
            'F': 8,    # News
            'G': 10,   # Method
            'H': 10,   # Phase
            'I': 10,   # Impact
            'J': 12,   # Sent.Score
            'K': 30,   # Key Phrases
            'L': 50,   # Event Reason Summary
            'M': 80,   # Full Analysis
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
    
    def create_portfolio_report(self, portfolio_data: List[Dict[str, Any]]) -> str:
        """
        Create a portfolio summary Excel report.
        
        Args:
            portfolio_data: List of stock analysis data dictionaries
            
        Returns:
            Path to generated Excel file
        """
        self.logger.info("Generating portfolio Excel report")
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create portfolio summary sheet
        ws = wb.create_sheet("Portfolio Summary")
        
        # Title
        ws['A1'] = "Portfolio Analysis Summary"
        ws['A1'].style = self.styles['title']
        ws.merge_cells('A1:P1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:P2')
        
        # Headers
        headers = [
            'Symbol', 'Company', 'Sector', 'Current Price', 'Day Change %',
            'Market Cap', '52W Range', 'Volume', 'RSI', 'Overall Signal',
            '1M Return %', '6M Return %', '1Y Return %', 'Volatility %',
            'Sharpe Ratio', 'Beta'
        ]
        
        row = 4
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
            ws.cell(row=row, column=col).style = self.styles['header']
        
        # Add portfolio data
        row = 5
        for stock_data in portfolio_data:
            summary = stock_data['summary_statistics']
            company_info = stock_data['company_info']
            signals = summary.get('signals', {})
            indicators = summary.get('technical_indicators', {})
            returns = summary.get('returns', {})
            
            # Basic info
            ws.cell(row=row, column=1, value=stock_data['ticker'])
            ws.cell(row=row, column=2, value=company_info.get('longName', 'N/A'))
            ws.cell(row=row, column=3, value=company_info.get('sector', 'N/A'))
            
            # Price data
            ws.cell(row=row, column=4, value=summary['current_price'])
            ws.cell(row=row, column=4).number_format = self.number_formats['currency']
            
            ws.cell(row=row, column=5, value=summary['price_change_1d_pct'] / 100)
            ws.cell(row=row, column=5).number_format = self.number_formats['percentage']
            
            # Color code price change
            if summary['price_change_1d_pct'] > 0:
                ws.cell(row=row, column=5).font = Font(color=self.colors['success'])
            elif summary['price_change_1d_pct'] < 0:
                ws.cell(row=row, column=5).font = Font(color=self.colors['danger'])
            
            # Market cap
            ws.cell(row=row, column=6, value=self.formatter.format_market_cap(company_info.get('marketCap', 0)))
            
            # 52-week range
            ws.cell(row=row, column=7, value=f"${summary['52_week_low']:.2f}-${summary['52_week_high']:.2f}")
            
            # Volume
            ws.cell(row=row, column=8, value=summary['volume'])
            ws.cell(row=row, column=8).number_format = self.number_formats['integer']
            
            # RSI
            rsi = indicators.get('rsi')
            if rsi:
                ws.cell(row=row, column=9, value=rsi)
                ws.cell(row=row, column=9).number_format = self.number_formats['number']
                
                # Color code RSI
                if rsi > 70:
                    ws.cell(row=row, column=9).font = Font(color=self.colors['danger'])
                elif rsi < 30:
                    ws.cell(row=row, column=9).font = Font(color=self.colors['success'])
            
            # Signal
            signal = signals.get('overall', 'Neutral')
            ws.cell(row=row, column=10, value=signal)
            if signal == 'Bullish':
                ws.cell(row=row, column=10).font = Font(color=self.colors['success'], bold=True)
            elif signal == 'Bearish':
                ws.cell(row=row, column=10).font = Font(color=self.colors['danger'], bold=True)
            
            # Returns
            ws.cell(row=row, column=11, value=returns.get('1_month', 0) / 100 if returns.get('1_month') else 0)
            ws.cell(row=row, column=12, value=returns.get('6_months', 0) / 100 if returns.get('6_months') else 0)
            ws.cell(row=row, column=13, value=returns.get('1_year', 0) / 100 if returns.get('1_year') else 0)
            
            for col in [11, 12, 13]:
                ws.cell(row=row, column=col).number_format = self.number_formats['percentage']
                value = ws.cell(row=row, column=col).value
                if value and value > 0:
                    ws.cell(row=row, column=col).font = Font(color=self.colors['success'])
                elif value and value < 0:
                    ws.cell(row=row, column=col).font = Font(color=self.colors['danger'])
            
            # Risk metrics
            ws.cell(row=row, column=14, value=summary.get('volatility_annual', 0) / 100)
            ws.cell(row=row, column=14).number_format = self.number_formats['percentage']
            
            ws.cell(row=row, column=15, value=summary.get('sharpe_ratio', 0))
            ws.cell(row=row, column=15).number_format = self.number_formats['number']
            
            ws.cell(row=row, column=16, value=summary.get('beta', 'N/A'))
            
            row += 1
        
        # Add summary statistics at the bottom
        row += 2
        ws[f'A{row}'] = "Portfolio Statistics"
        ws[f'A{row}'].style = self.styles['subheader']
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        # Calculate portfolio metrics
        total_stocks = len(portfolio_data)
        bullish_count = sum(1 for d in portfolio_data if d['summary_statistics'].get('signals', {}).get('overall') == 'Bullish')
        bearish_count = sum(1 for d in portfolio_data if d['summary_statistics'].get('signals', {}).get('overall') == 'Bearish')
        
        stats = [
            ['Total Stocks', total_stocks],
            ['Bullish Signals', bullish_count],
            ['Bearish Signals', bearish_count],
            ['Neutral Signals', total_stocks - bullish_count - bearish_count]
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            # Find the first non-merged cell to get the column letter
            column_letter = None
            for cell in column_cells:
                if not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter
                    break
            if not column_letter:
                continue  # skip if all cells are merged

            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save report
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"portfolio_analysis_{timestamp}.xlsx"
        filepath = config.OUTPUT_DIR / filename
        
        wb.save(filepath)
        self.logger.info(f"Portfolio report saved to {filepath}")
        
        return str(filepath)